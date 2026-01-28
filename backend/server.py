from fastapi import FastAPI, APIRouter, HTTPException, Depends, Response, UploadFile, File, BackgroundTasks
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
import base64
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
import resend

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Resend configuration
resend.api_key = os.environ.get('RESEND_API_KEY', '')
ALERT_EMAIL = os.environ.get('ALERT_EMAIL', '')
ALERT_DAYS_BEFORE = int(os.environ.get('ALERT_DAYS_BEFORE', 7))
SENDER_EMAIL = "onboarding@resend.dev"

app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

JWT_SECRET = os.environ.get('JWT_SECRET', 'warehouse-construction-secret-key-2024')
JWT_ALGORITHM = 'HS256'

# Create uploads directory
UPLOAD_DIR = ROOT_DIR / 'uploads'
UPLOAD_DIR.mkdir(exist_ok=True)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== AUTH MODELS ====================
class UserCreate(BaseModel):
    name: str
    email: EmailStr
    password: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    email: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

# ==================== EQUIPAMENTO MODEL ====================
class EquipamentoCreate(BaseModel):
    codigo: str
    descricao: str
    marca: str = ""
    modelo: str = ""
    data_aquisicao: Optional[str] = None
    ativo: bool = True
    categoria: str = ""
    numero_serie: str = ""
    responsavel: str = ""
    estado_conservacao: str = "Bom"
    foto: str = ""
    local_id: Optional[str] = None

class Equipamento(EquipamentoCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tipo: str = "Equipamento"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== VIATURA MODEL ====================
class ViaturaCreate(BaseModel):
    matricula: str
    marca: str = ""
    modelo: str = ""
    combustivel: str = "Gasoleo"
    ativa: bool = True
    foto: str = ""
    data_vistoria: Optional[str] = None
    data_seguro: Optional[str] = None
    documento_unico: str = ""
    apolice_seguro: str = ""
    observacoes: str = ""
    local_id: Optional[str] = None

class Viatura(ViaturaCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== MATERIAL MODEL ====================
class MaterialCreate(BaseModel):
    codigo: str
    descricao: str
    unidade: str = "unidade"
    stock_atual: float = 0
    stock_minimo: float = 0
    ativo: bool = True
    local_id: Optional[str] = None

class Material(MaterialCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== LOCAL MODEL ====================
class LocalCreate(BaseModel):
    codigo: str
    nome: str
    tipo: str = "ARM"
    obra_id: Optional[str] = None
    ativo: bool = True

class Local(LocalCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== OBRA MODEL ====================
class ObraCreate(BaseModel):
    codigo: str
    nome: str
    endereco: str = ""
    cliente: str = ""
    estado: str = "Ativa"

class Obra(ObraCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== MOVIMENTO MODELS ====================
class MovimentoAtivoCreate(BaseModel):
    ativo_id: str
    tipo_ativo: str = "equipamento"
    tipo_movimento: str
    origem_id: Optional[str] = None
    destino_id: Optional[str] = None
    responsavel: str = ""
    observacoes: str = ""

class MovimentoAtivo(MovimentoAtivoCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    data_hora: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class MovimentoStockCreate(BaseModel):
    material_id: str
    tipo_movimento: str
    quantidade: float
    obra_id: Optional[str] = None
    fornecedor: str = ""
    documento: str = ""
    responsavel: str = ""
    observacoes: str = ""

class MovimentoStock(MovimentoStockCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    data_hora: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class MovimentoViaturaCreate(BaseModel):
    viatura_id: str
    obra_id: Optional[str] = None
    condutor: str = ""
    km_inicial: float = 0
    km_final: float = 0
    data: str = ""
    observacoes: str = ""

class MovimentoViatura(MovimentoViaturaCreate):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== EMAIL MODEL ====================
class EmailAlert(BaseModel):
    viatura_id: str
    matricula: str
    marca: str
    modelo: str
    tipo_alerta: str  # "vistoria" or "seguro"
    data_expiracao: str
    dias_restantes: int

# ==================== AUTH FUNCTIONS ====================
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=7)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ==================== EMAIL FUNCTIONS ====================
async def send_alert_email(alerts: List[EmailAlert]):
    if not ALERT_EMAIL or not resend.api_key:
        logger.warning("Email alerts not configured")
        return
    
    if not alerts:
        return
    
    # Build HTML email
    alerts_html = ""
    for alert in alerts:
        tipo = "Vistoria" if alert.tipo_alerta == "vistoria" else "Seguro"
        urgency_class = "color: #dc2626;" if alert.dias_restantes <= 0 else "color: #d97706;"
        status = "EXPIRADO" if alert.dias_restantes <= 0 else f"Expira em {alert.dias_restantes} dias"
        
        alerts_html += f"""
        <tr>
            <td style="padding: 12px; border-bottom: 1px solid #e5e7eb;">{alert.matricula}</td>
            <td style="padding: 12px; border-bottom: 1px solid #e5e7eb;">{alert.marca} {alert.modelo}</td>
            <td style="padding: 12px; border-bottom: 1px solid #e5e7eb;">{tipo}</td>
            <td style="padding: 12px; border-bottom: 1px solid #e5e7eb;">{alert.data_expiracao}</td>
            <td style="padding: 12px; border-bottom: 1px solid #e5e7eb; {urgency_class} font-weight: bold;">{status}</td>
        </tr>
        """
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
    </head>
    <body style="font-family: Arial, sans-serif; margin: 0; padding: 20px; background-color: #f8fafc;">
        <div style="max-width: 600px; margin: 0 auto; background-color: white; border-radius: 8px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.1);">
            <div style="background-color: #1e293b; color: white; padding: 20px; text-align: center;">
                <h1 style="margin: 0; font-size: 24px;">⚠️ Alertas de Viaturas</h1>
                <p style="margin: 10px 0 0 0; opacity: 0.8;">Gestão de Armazém - Construção Civil</p>
            </div>
            <div style="padding: 20px;">
                <p style="color: #475569; margin-bottom: 20px;">
                    As seguintes viaturas têm vistorias ou seguros a expirar nos próximos {ALERT_DAYS_BEFORE} dias:
                </p>
                <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
                    <thead>
                        <tr style="background-color: #f1f5f9;">
                            <th style="padding: 12px; text-align: left; font-weight: 600;">Matrícula</th>
                            <th style="padding: 12px; text-align: left; font-weight: 600;">Viatura</th>
                            <th style="padding: 12px; text-align: left; font-weight: 600;">Tipo</th>
                            <th style="padding: 12px; text-align: left; font-weight: 600;">Data</th>
                            <th style="padding: 12px; text-align: left; font-weight: 600;">Estado</th>
                        </tr>
                    </thead>
                    <tbody>
                        {alerts_html}
                    </tbody>
                </table>
                <p style="color: #64748b; font-size: 12px; margin-top: 20px; padding-top: 20px; border-top: 1px solid #e5e7eb;">
                    Este email foi enviado automaticamente pelo sistema de Gestão de Armazém.
                </p>
            </div>
        </div>
    </body>
    </html>
    """
    
    params = {
        "from": SENDER_EMAIL,
        "to": [ALERT_EMAIL],
        "subject": f"⚠️ Alertas de Viaturas - {len(alerts)} alerta(s)",
        "html": html_content
    }
    
    try:
        email = await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Alert email sent successfully: {email.get('id')}")
        return email
    except Exception as e:
        logger.error(f"Failed to send alert email: {str(e)}")
        raise

async def check_and_send_alerts():
    """Check for expiring vistorias and seguros and send email alerts"""
    viaturas = await db.viaturas.find({"ativa": True}, {"_id": 0}).to_list(1000)
    today = datetime.now(timezone.utc).date()
    alerts = []
    
    for v in viaturas:
        # Check vistoria
        if v.get("data_vistoria"):
            try:
                vistoria_date = datetime.fromisoformat(v["data_vistoria"].replace("Z", "+00:00")).date()
                days_until = (vistoria_date - today).days
                if days_until <= ALERT_DAYS_BEFORE:
                    alerts.append(EmailAlert(
                        viatura_id=v["id"],
                        matricula=v["matricula"],
                        marca=v.get("marca", ""),
                        modelo=v.get("modelo", ""),
                        tipo_alerta="vistoria",
                        data_expiracao=vistoria_date.strftime("%d/%m/%Y"),
                        dias_restantes=days_until
                    ))
            except:
                pass
        
        # Check seguro
        if v.get("data_seguro"):
            try:
                seguro_date = datetime.fromisoformat(v["data_seguro"].replace("Z", "+00:00")).date()
                days_until = (seguro_date - today).days
                if days_until <= ALERT_DAYS_BEFORE:
                    alerts.append(EmailAlert(
                        viatura_id=v["id"],
                        matricula=v["matricula"],
                        marca=v.get("marca", ""),
                        modelo=v.get("modelo", ""),
                        tipo_alerta="seguro",
                        data_expiracao=seguro_date.strftime("%d/%m/%Y"),
                        dias_restantes=days_until
                    ))
            except:
                pass
    
    if alerts:
        await send_alert_email(alerts)
    
    return alerts

# ==================== AUTH ROUTES ====================
@api_router.post("/auth/register", response_model=TokenResponse)
async def register(data: UserCreate):
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "name": data.name,
        "email": data.email,
        "password": hash_password(data.password),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    token = create_token(user_id)
    return TokenResponse(
        access_token=token,
        user=UserResponse(id=user_id, name=data.name, email=data.email)
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(data: UserLogin):
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user or not verify_password(data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user["id"])
    return TokenResponse(
        access_token=token,
        user=UserResponse(id=user["id"], name=user["name"], email=user["email"])
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user=Depends(get_current_user)):
    return UserResponse(id=user["id"], name=user["name"], email=user["email"])

# ==================== UPLOAD ROUTES ====================
@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), user=Depends(get_current_user)):
    """Upload an image file and return the URL"""
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Only image files are allowed")
    
    # Generate unique filename
    ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    filename = f"{uuid.uuid4()}.{ext}"
    filepath = UPLOAD_DIR / filename
    
    # Save file
    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)
    
    # Return URL (relative path that can be served)
    return {"url": f"/api/uploads/{filename}", "filename": filename}

@api_router.get("/uploads/{filename}")
async def get_upload(filename: str):
    """Serve uploaded files"""
    filepath = UPLOAD_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    # Determine content type
    ext = filename.split(".")[-1].lower()
    content_types = {
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png",
        "gif": "image/gif",
        "webp": "image/webp"
    }
    content_type = content_types.get(ext, "application/octet-stream")
    
    with open(filepath, "rb") as f:
        content = f.read()
    
    return Response(content=content, media_type=content_type)

# ==================== EQUIPAMENTO ROUTES ====================
@api_router.get("/equipamentos", response_model=List[Equipamento])
async def get_equipamentos(user=Depends(get_current_user)):
    items = await db.equipamentos.find({}, {"_id": 0}).to_list(1000)
    return items

@api_router.post("/equipamentos", response_model=Equipamento)
async def create_equipamento(data: EquipamentoCreate, user=Depends(get_current_user)):
    existing = await db.equipamentos.find_one({"codigo": data.codigo}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Código já existe")
    
    equipamento = Equipamento(**data.model_dump())
    doc = equipamento.model_dump()
    await db.equipamentos.insert_one(doc)
    return equipamento

@api_router.put("/equipamentos/{equipamento_id}", response_model=Equipamento)
async def update_equipamento(equipamento_id: str, data: EquipamentoCreate, user=Depends(get_current_user)):
    existing = await db.equipamentos.find_one({"id": equipamento_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Equipamento não encontrado")
    
    update_data = data.model_dump()
    await db.equipamentos.update_one({"id": equipamento_id}, {"$set": update_data})
    updated = await db.equipamentos.find_one({"id": equipamento_id}, {"_id": 0})
    return updated

@api_router.delete("/equipamentos/{equipamento_id}")
async def delete_equipamento(equipamento_id: str, user=Depends(get_current_user)):
    result = await db.equipamentos.delete_one({"id": equipamento_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Equipamento não encontrado")
    return {"message": "Equipamento eliminado"}

# ==================== VIATURA ROUTES ====================
@api_router.get("/viaturas", response_model=List[Viatura])
async def get_viaturas(user=Depends(get_current_user)):
    items = await db.viaturas.find({}, {"_id": 0}).to_list(1000)
    return items

@api_router.post("/viaturas", response_model=Viatura)
async def create_viatura(data: ViaturaCreate, user=Depends(get_current_user)):
    existing = await db.viaturas.find_one({"matricula": data.matricula}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Matrícula já existe")
    
    viatura = Viatura(**data.model_dump())
    doc = viatura.model_dump()
    await db.viaturas.insert_one(doc)
    return viatura

@api_router.put("/viaturas/{viatura_id}", response_model=Viatura)
async def update_viatura(viatura_id: str, data: ViaturaCreate, user=Depends(get_current_user)):
    existing = await db.viaturas.find_one({"id": viatura_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Viatura não encontrada")
    
    update_data = data.model_dump()
    await db.viaturas.update_one({"id": viatura_id}, {"$set": update_data})
    updated = await db.viaturas.find_one({"id": viatura_id}, {"_id": 0})
    return updated

@api_router.delete("/viaturas/{viatura_id}")
async def delete_viatura(viatura_id: str, user=Depends(get_current_user)):
    result = await db.viaturas.delete_one({"id": viatura_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Viatura não encontrada")
    return {"message": "Viatura eliminada"}

# ==================== MATERIAL ROUTES ====================
@api_router.get("/materiais", response_model=List[Material])
async def get_materiais(user=Depends(get_current_user)):
    items = await db.materiais.find({}, {"_id": 0}).to_list(1000)
    return items

@api_router.post("/materiais", response_model=Material)
async def create_material(data: MaterialCreate, user=Depends(get_current_user)):
    existing = await db.materiais.find_one({"codigo": data.codigo}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Código já existe")
    
    material = Material(**data.model_dump())
    doc = material.model_dump()
    await db.materiais.insert_one(doc)
    return material

@api_router.put("/materiais/{material_id}", response_model=Material)
async def update_material(material_id: str, data: MaterialCreate, user=Depends(get_current_user)):
    existing = await db.materiais.find_one({"id": material_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Material não encontrado")
    
    update_data = data.model_dump()
    await db.materiais.update_one({"id": material_id}, {"$set": update_data})
    updated = await db.materiais.find_one({"id": material_id}, {"_id": 0})
    return updated

@api_router.delete("/materiais/{material_id}")
async def delete_material(material_id: str, user=Depends(get_current_user)):
    result = await db.materiais.delete_one({"id": material_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Material não encontrado")
    return {"message": "Material eliminado"}

# ==================== LOCAL ROUTES ====================
@api_router.get("/locais", response_model=List[Local])
async def get_locais(user=Depends(get_current_user)):
    items = await db.locais.find({}, {"_id": 0}).to_list(1000)
    return items

@api_router.post("/locais", response_model=Local)
async def create_local(data: LocalCreate, user=Depends(get_current_user)):
    existing = await db.locais.find_one({"codigo": data.codigo}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Código já existe")
    
    local = Local(**data.model_dump())
    doc = local.model_dump()
    await db.locais.insert_one(doc)
    return local

@api_router.put("/locais/{local_id}", response_model=Local)
async def update_local(local_id: str, data: LocalCreate, user=Depends(get_current_user)):
    existing = await db.locais.find_one({"id": local_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Local não encontrado")
    
    update_data = data.model_dump()
    await db.locais.update_one({"id": local_id}, {"$set": update_data})
    updated = await db.locais.find_one({"id": local_id}, {"_id": 0})
    return updated

@api_router.delete("/locais/{local_id}")
async def delete_local(local_id: str, user=Depends(get_current_user)):
    result = await db.locais.delete_one({"id": local_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Local não encontrado")
    return {"message": "Local eliminado"}

# ==================== OBRA ROUTES ====================
@api_router.get("/obras", response_model=List[Obra])
async def get_obras(user=Depends(get_current_user)):
    items = await db.obras.find({}, {"_id": 0}).to_list(1000)
    return items

@api_router.post("/obras", response_model=Obra)
async def create_obra(data: ObraCreate, user=Depends(get_current_user)):
    existing = await db.obras.find_one({"codigo": data.codigo}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Código já existe")
    
    obra = Obra(**data.model_dump())
    doc = obra.model_dump()
    await db.obras.insert_one(doc)
    return obra

@api_router.put("/obras/{obra_id}", response_model=Obra)
async def update_obra(obra_id: str, data: ObraCreate, user=Depends(get_current_user)):
    existing = await db.obras.find_one({"id": obra_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    
    update_data = data.model_dump()
    await db.obras.update_one({"id": obra_id}, {"$set": update_data})
    updated = await db.obras.find_one({"id": obra_id}, {"_id": 0})
    return updated

@api_router.delete("/obras/{obra_id}")
async def delete_obra(obra_id: str, user=Depends(get_current_user)):
    result = await db.obras.delete_one({"id": obra_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    await db.locais.update_many({"obra_id": obra_id}, {"$set": {"obra_id": None}})
    return {"message": "Obra eliminada"}

@api_router.get("/obras/{obra_id}/recursos")
async def get_obra_recursos(obra_id: str, user=Depends(get_current_user)):
    obra = await db.obras.find_one({"id": obra_id}, {"_id": 0})
    if not obra:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    
    locais = await db.locais.find({"obra_id": obra_id}, {"_id": 0}).to_list(1000)
    local_ids = [l["id"] for l in locais]
    
    equipamentos = await db.equipamentos.find({"local_id": {"$in": local_ids}}, {"_id": 0}).to_list(1000)
    viaturas = await db.viaturas.find({"local_id": {"$in": local_ids}}, {"_id": 0}).to_list(1000)
    materiais = await db.materiais.find({"local_id": {"$in": local_ids}}, {"_id": 0}).to_list(1000)
    
    return {
        "obra": obra,
        "locais": locais,
        "equipamentos": equipamentos,
        "viaturas": viaturas,
        "materiais": materiais
    }

# ==================== MOVIMENTO ROUTES ====================
@api_router.get("/movimentos/ativos", response_model=List[MovimentoAtivo])
async def get_movimentos_ativos(user=Depends(get_current_user)):
    items = await db.movimentos_ativos.find({}, {"_id": 0}).to_list(1000)
    return items

@api_router.post("/movimentos/ativos", response_model=MovimentoAtivo)
async def create_movimento_ativo(data: MovimentoAtivoCreate, user=Depends(get_current_user)):
    movimento = MovimentoAtivo(**data.model_dump())
    doc = movimento.model_dump()
    await db.movimentos_ativos.insert_one(doc)
    
    if data.destino_id:
        await db.equipamentos.update_one(
            {"id": data.ativo_id},
            {"$set": {"local_id": data.destino_id}}
        )
    
    return movimento

@api_router.get("/movimentos/stock", response_model=List[MovimentoStock])
async def get_movimentos_stock(user=Depends(get_current_user)):
    items = await db.movimentos_stock.find({}, {"_id": 0}).to_list(1000)
    return items

@api_router.post("/movimentos/stock", response_model=MovimentoStock)
async def create_movimento_stock(data: MovimentoStockCreate, user=Depends(get_current_user)):
    movimento = MovimentoStock(**data.model_dump())
    doc = movimento.model_dump()
    await db.movimentos_stock.insert_one(doc)
    
    material = await db.materiais.find_one({"id": data.material_id}, {"_id": 0})
    if material:
        new_stock = material.get("stock_atual", 0)
        if data.tipo_movimento == "Entrada":
            new_stock += data.quantidade
        else:
            new_stock -= data.quantidade
        
        await db.materiais.update_one(
            {"id": data.material_id},
            {"$set": {"stock_atual": new_stock}}
        )
    
    return movimento

@api_router.get("/movimentos/viaturas", response_model=List[MovimentoViatura])
async def get_movimentos_viaturas(user=Depends(get_current_user)):
    items = await db.movimentos_viaturas.find({}, {"_id": 0}).to_list(1000)
    return items

@api_router.post("/movimentos/viaturas", response_model=MovimentoViatura)
async def create_movimento_viatura(data: MovimentoViaturaCreate, user=Depends(get_current_user)):
    movimento = MovimentoViatura(**data.model_dump())
    doc = movimento.model_dump()
    await db.movimentos_viaturas.insert_one(doc)
    return movimento

# ==================== ALERTS ROUTES ====================
@api_router.get("/alerts/check")
async def check_alerts(user=Depends(get_current_user)):
    """Check for expiring vistorias and seguros"""
    viaturas = await db.viaturas.find({"ativa": True}, {"_id": 0}).to_list(1000)
    today = datetime.now(timezone.utc).date()
    alerts = []
    
    for v in viaturas:
        if v.get("data_vistoria"):
            try:
                vistoria_date = datetime.fromisoformat(v["data_vistoria"].replace("Z", "+00:00")).date()
                days_until = (vistoria_date - today).days
                if days_until <= ALERT_DAYS_BEFORE:
                    alerts.append({
                        "viatura_id": v["id"],
                        "matricula": v["matricula"],
                        "marca": v.get("marca", ""),
                        "modelo": v.get("modelo", ""),
                        "tipo_alerta": "vistoria",
                        "data_expiracao": vistoria_date.strftime("%d/%m/%Y"),
                        "dias_restantes": days_until
                    })
            except:
                pass
        
        if v.get("data_seguro"):
            try:
                seguro_date = datetime.fromisoformat(v["data_seguro"].replace("Z", "+00:00")).date()
                days_until = (seguro_date - today).days
                if days_until <= ALERT_DAYS_BEFORE:
                    alerts.append({
                        "viatura_id": v["id"],
                        "matricula": v["matricula"],
                        "marca": v.get("marca", ""),
                        "modelo": v.get("modelo", ""),
                        "tipo_alerta": "seguro",
                        "data_expiracao": seguro_date.strftime("%d/%m/%Y"),
                        "dias_restantes": days_until
                    })
            except:
                pass
    
    return {"alerts": alerts, "total": len(alerts)}

@api_router.post("/alerts/send")
async def send_alerts(background_tasks: BackgroundTasks, user=Depends(get_current_user)):
    """Send email alerts for expiring vistorias and seguros"""
    try:
        alerts = await check_and_send_alerts()
        return {
            "status": "success",
            "message": f"Email enviado com {len(alerts)} alerta(s)" if alerts else "Não há alertas para enviar",
            "alerts_count": len(alerts)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao enviar email: {str(e)}")

# ==================== SUMMARY ROUTE ====================
@api_router.get("/summary")
async def get_summary(user=Depends(get_current_user)):
    equipamentos = await db.equipamentos.find({}, {"_id": 0}).to_list(1000)
    viaturas = await db.viaturas.find({}, {"_id": 0}).to_list(1000)
    materiais = await db.materiais.find({}, {"_id": 0}).to_list(1000)
    locais = await db.locais.find({}, {"_id": 0}).to_list(1000)
    obras = await db.obras.find({}, {"_id": 0}).to_list(1000)
    
    alerts = []
    today = datetime.now(timezone.utc).date()
    
    for v in viaturas:
        if v.get("data_vistoria"):
            try:
                vistoria_date = datetime.fromisoformat(v["data_vistoria"].replace("Z", "+00:00")).date()
                days_until = (vistoria_date - today).days
                if days_until <= ALERT_DAYS_BEFORE:
                    alerts.append({
                        "type": "vistoria",
                        "item": f"{v['marca']} {v['modelo']} ({v['matricula']})",
                        "message": f"Vistoria em {days_until} dias" if days_until >= 0 else "Vistoria expirada",
                        "urgent": days_until < 0
                    })
            except:
                pass
        
        if v.get("data_seguro"):
            try:
                seguro_date = datetime.fromisoformat(v["data_seguro"].replace("Z", "+00:00")).date()
                days_until = (seguro_date - today).days
                if days_until <= ALERT_DAYS_BEFORE:
                    alerts.append({
                        "type": "seguro",
                        "item": f"{v['marca']} {v['modelo']} ({v['matricula']})",
                        "message": f"Seguro expira em {days_until} dias" if days_until >= 0 else "Seguro expirado",
                        "urgent": days_until < 0
                    })
            except:
                pass
    
    for m in materiais:
        if m.get("stock_atual", 0) <= m.get("stock_minimo", 0) and m.get("stock_minimo", 0) > 0:
            alerts.append({
                "type": "stock",
                "item": f"{m['codigo']} - {m['descricao']}",
                "message": f"Stock baixo: {m.get('stock_atual', 0)} {m.get('unidade', 'un')} (mín: {m.get('stock_minimo', 0)})",
                "urgent": m.get("stock_atual", 0) == 0
            })
    
    return {
        "equipamentos": {
            "total": len(equipamentos),
            "ativos": len([e for e in equipamentos if e.get("ativo", True)]),
            "inativos": len([e for e in equipamentos if not e.get("ativo", True)])
        },
        "viaturas": {
            "total": len(viaturas),
            "ativas": len([v for v in viaturas if v.get("ativa", True)]),
            "inativas": len([v for v in viaturas if not v.get("ativa", True)])
        },
        "materiais": {
            "total": len(materiais),
            "stock_total": sum(m.get("stock_atual", 0) for m in materiais)
        },
        "locais": {
            "total": len(locais),
            "armazens": len([l for l in locais if l.get("tipo") == "ARM"]),
            "oficinas": len([l for l in locais if l.get("tipo") == "OFI"]),
            "obras": len([l for l in locais if l.get("tipo") == "OBR"])
        },
        "obras": {
            "total": len(obras),
            "ativas": len([o for o in obras if o.get("estado") == "Ativa"]),
            "concluidas": len([o for o in obras if o.get("estado") == "Concluida"]),
            "pausadas": len([o for o in obras if o.get("estado") == "Pausada"])
        },
        "alerts": alerts
    }

# ==================== EXPORT ROUTES ====================
@api_router.get("/export/pdf")
async def export_pdf(user=Depends(get_current_user)):
    equipamentos = await db.equipamentos.find({}, {"_id": 0}).to_list(1000)
    viaturas = await db.viaturas.find({}, {"_id": 0}).to_list(1000)
    materiais = await db.materiais.find({}, {"_id": 0}).to_list(1000)
    obras = await db.obras.find({}, {"_id": 0}).to_list(1000)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    
    elements.append(Paragraph("Relatório de Armazém - Construção Civil", styles['Title']))
    elements.append(Paragraph(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    summary_data = [
        ["Categoria", "Total", "Ativos", "Inativos"],
        ["Equipamentos", len(equipamentos), len([e for e in equipamentos if e.get("ativo")]), len([e for e in equipamentos if not e.get("ativo")])],
        ["Viaturas", len(viaturas), len([v for v in viaturas if v.get("ativa")]), len([v for v in viaturas if not v.get("ativa")])],
        ["Materiais", len(materiais), "-", "-"],
        ["Obras", len(obras), len([o for o in obras if o.get("estado") == "Ativa"]), len([o for o in obras if o.get("estado") != "Ativa"])]
    ]
    
    table = Table(summary_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=relatorio_armazem.pdf"}
    )

@api_router.get("/export/excel")
async def export_excel(user=Depends(get_current_user)):
    equipamentos = await db.equipamentos.find({}, {"_id": 0}).to_list(1000)
    viaturas = await db.viaturas.find({}, {"_id": 0}).to_list(1000)
    materiais = await db.materiais.find({}, {"_id": 0}).to_list(1000)
    locais = await db.locais.find({}, {"_id": 0}).to_list(1000)
    obras = await db.obras.find({}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    
    ws = wb.active
    ws.title = "Equipamentos"
    ws.append(["Código", "Descrição", "Marca", "Modelo", "Categoria", "Nº Série", "Estado Conservação", "Responsável", "Ativo"])
    for e in equipamentos:
        ws.append([e.get("codigo", ""), e.get("descricao", ""), e.get("marca", ""), e.get("modelo", ""), 
                   e.get("categoria", ""), e.get("numero_serie", ""), e.get("estado_conservacao", ""),
                   e.get("responsavel", ""), "Sim" if e.get("ativo") else "Não"])
    
    ws = wb.create_sheet("Viaturas")
    ws.append(["Matrícula", "Marca", "Modelo", "Combustível", "Data Vistoria", "Data Seguro", "Apólice", "Ativa", "Observações"])
    for v in viaturas:
        ws.append([v.get("matricula", ""), v.get("marca", ""), v.get("modelo", ""), v.get("combustivel", ""),
                   v.get("data_vistoria", ""), v.get("data_seguro", ""), v.get("apolice_seguro", ""),
                   "Sim" if v.get("ativa") else "Não", v.get("observacoes", "")])
    
    ws = wb.create_sheet("Materiais")
    ws.append(["Código", "Descrição", "Unidade", "Stock Atual", "Stock Mínimo", "Ativo"])
    for m in materiais:
        ws.append([m.get("codigo", ""), m.get("descricao", ""), m.get("unidade", ""),
                   m.get("stock_atual", 0), m.get("stock_minimo", 0), "Sim" if m.get("ativo") else "Não"])
    
    ws = wb.create_sheet("Locais")
    ws.append(["Código", "Nome", "Tipo", "Ativo"])
    for l in locais:
        ws.append([l.get("codigo", ""), l.get("nome", ""), l.get("tipo", ""), "Sim" if l.get("ativo") else "Não"])
    
    ws = wb.create_sheet("Obras")
    ws.append(["Código", "Nome", "Endereço", "Cliente", "Estado"])
    for o in obras:
        ws.append([o.get("codigo", ""), o.get("nome", ""), o.get("endereco", ""), o.get("cliente", ""), o.get("estado", "")])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio_armazem.xlsx"}
    )

# ==================== ROOT ====================
@api_router.get("/")
async def root():
    return {"message": "Warehouse Management API"}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
